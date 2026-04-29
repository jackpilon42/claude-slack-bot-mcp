#!/usr/bin/env python3
"""
Scan an .xlsx for formatting that differs from a "plain blank" cell, and emit
plain-English things to say to the Slack bot (one sentence per line).

Phrasing targets Claude → Smartsheet (column numbers + row numbers, not Excel A1).

Sections emitted:
  - Bold: non-empty text + font bold.
  - Background: solid fill not white/near-white/light-gray.
  - Merged ranges (one command per merge).
  - Row height: uniform explicit height across the used grid.
  - Column width: unusually narrow or wide columns (sample).
  - Bottom borders: consecutive runs (per row) of the same bottom border style.
  - Text wrap: horizontal runs of wrap=True per row.
  - Number formats: vertical runs of the same non-General format per column.
  - Alignment: horizontal center on early header rows (vertical is omitted — openpyxl often defaults vertical to center).
  - Headline font: large blue text on row 1; white text runs on row 1.

Usage:
  python3 scripts/generate_excel_format_bot_commands.py [path/to.xlsx] > scripts/project_timeline_bot_commands.txt
"""
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    raise

LIGHT_FILLS = {"#FFFFFF", "#EFEFEF", "#F8F8F8", "#D9D9D9", "#FFFFFE"}

BOLD_TEMPLATES = [
    "Bold column {c} row {r}.",
    "Make column {c} row {r} bold.",
    "Set column {c} row {r} to bold.",
    "Column {c} row {r}: bold.",
    "At column {c} row {r}, use bold.",
    "Apply bold to column {c} row {r}.",
    "Column {c} row {r} — bold.",
    "Row {r}, column {c}: bold.",
    "Put bold on column {c} row {r}.",
    "Use bold at column {c} row {r}.",
    "Column {c} row {r} should render bold.",
    "Emphasize column {c} row {r}.",
    "Emphasize column {c} row {r} like the template.",
    "Template has column {c} row {r} bold; mirror that.",
    "For row {r}, column {c}, make that cell bold.",
    "Please bold column {c} row {r}.",
    "Could you bold column {c} row {r}?",
    "I'd like column {c} row {r} in bold.",
    "Can you put column {c} row {r} in bold?",
    "Turn on bold for column {c} row {r}.",
]

HEX_TO_COLOR_WORD = {
    "#434343": "gray",
    "#0B5394": "blue",
    "#45818E": "green",
    "#B85B22": "orange",
    "#38761D": "green",
    "#351C75": "purple",
}

_NAMED_RGB = {
    "red": (255, 0, 0),
    "green": (0, 255, 0),
    "blue": (0, 0, 255),
    "yellow": (255, 255, 0),
    "orange": (255, 165, 0),
    "white": (255, 255, 255),
    "black": (0, 0, 0),
    "purple": (128, 0, 128),
    "pink": (255, 192, 203),
    "gray": (128, 128, 128),
    "grey": (128, 128, 128),
    "brown": (139, 69, 19),
}

FILL_TEMPLATES = [
    "For row {r} in column {c}, use a {color} background.",
    "Highlight column {c} row {r} with {color}.",
    "Please set column {c} on row {r} to a {color} fill.",
    "Column {c} on row {r} should be {color}.",
    "On column {c}, row {r} should have a {color} background.",
    "Could you make column {c} row {r} {color}?",
    "Turn column {c} row {r} {color}.",
    "I'd like column {c} on row {r} filled in {color}.",
    "Give column {c} row {r} a {color} background, please.",
    "For column {c}, row {r} needs a {color} fill.",
]

MERGE_SINGLE_ROW_TEMPLATES = [
    "Merge columns {c1} through {c2} on row {r}.",
    "On row {r}, merge columns {c1} to {c2} into one cell.",
    "Combine columns {c1} through {c2} on row {r}.",
    "Span columns {c1}–{c2} on row {r} as a single merged cell.",
]

MERGE_MULTI_ROW_TEMPLATES = [
    "Merge columns {c1} through {c2} from row {r1} to row {r2}.",
    "Combine columns {c1} to {c2} across rows {r1} through {r2}.",
    "Use one merged block for columns {c1}–{c2} rows {r1}–{r2}.",
]

MERGE_MULTI_ONE_COL_TEMPLATES = [
    "Merge column {c} from row {r1} down to row {r2}.",
    "Span column {c} across rows {r1} through {r2} as one cell.",
]

WRAP_TEMPLATES = [
    "Wrap text in columns {c1} through {c2} on row {r}.",
    "Turn on text wrapping for columns {c1} to {c2} on row {r}.",
    "For row {r}, enable wrap for columns {c1} through {c2}.",
]

WRAP_ONE_COL_TEMPLATES = [
    "Wrap the text in column {c} row {r}.",
    "Enable wrapping on column {c} row {r}.",
]

BORDER_BOTTOM_TEMPLATES = [
    "Add a {style} bottom border to row {r} from column {c1} to column {c2}.",
    "Put a {style} line along the bottom of row {r} across columns {c1} through {c2}.",
    "Underline row {r} with a {style} border from column {c1} to column {c2}.",
]

NUMFMT_RANGE_TEMPLATES = [
    "Format column {c} rows {r1} through {r2} as {desc}.",
    "From row {r1} to row {r2} in column {c}, use {desc} formatting.",
]

NUMFMT_ONE_TEMPLATES = [
    "Format column {c} row {r} as {desc}.",
    "On column {c} row {r}, display values as {desc}.",
]

ROW_HEIGHT_UNIFORM_TEMPLATES = [
    "Set rows {r1} through {r2} to a uniform height of about {h} points.",
    "Make row heights {r1}–{r2} roughly {h} points each.",
]

COL_WIDTH_WIDE_TEMPLATES = [
    "Set column {c} to a width of about {w} characters.",
    "Widen column {c} to roughly {w} characters wide.",
]

COL_WIDTH_NARROW_TEMPLATES = [
    "Narrow column {c} to about {w} characters wide.",
    "Set column {c} to a tight width near {w} characters.",
]

ALIGN_H_CENTER_TEMPLATES = [
    "Center-align columns {c1} through {c2} on row {r}.",
    "Horizontally center the text in columns {c1} to {c2} on row {r}.",
]

HEADLINE_BLUE_TEMPLATES = [
    "Use a large blue heading on column {c} row {r} (about {sz} point).",
    "Set column {c} row {r} to a big blue title near {sz} point.",
]

WHITE_FONT_TEMPLATES = [
    "Set the font color to white for columns {c1} through {c2} on row {r}.",
    "Use white text in columns {c1} to {c2} on row {r}.",
]

EXAMPLE_MULTI_ROW_RANGE = "Highlight column 2 rows 8 through 10 with gray."

BOTTOM_STYLE_WORD = {
    "thick": "thick",
    "thin": "thin",
    "medium": "medium",
    "hair": "hairline",
    "double": "double-line",
}


def hx_from_color(c):
    if c is None:
        return None
    r = getattr(c, "rgb", None)
    if r is None:
        return None
    s = str(r).upper()
    if len(s) == 8 and s.startswith("FF"):
        return "#" + s[2:8]
    return None


def _hex_to_rgb(hx):
    h = hx.strip().lstrip("#").upper()
    if len(h) != 6:
        return None
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def nearest_color_word(hx):
    u = hx.strip().upper()
    if u in HEX_TO_COLOR_WORD:
        return HEX_TO_COLOR_WORD[u]
    rgb = _hex_to_rgb(u)
    if not rgb:
        return "gray"
    best_name = "gray"
    best_d = None
    for name, ref in _NAMED_RGB.items():
        if name == "grey":
            continue
        d = sum((a - b) ** 2 for a, b in zip(rgb, ref))
        if best_d is None or d < best_d:
            best_d, best_name = d, name
    return best_name


def bottom_border_style(cell):
    b = cell.border
    if not b or not b.bottom:
        return None
    return b.bottom.style


def _consecutive_runs(indices):
    if not indices:
        return []
    s = e = indices[0]
    out = []
    for x in indices[1:]:
        if x == e + 1:
            e = x
        else:
            out.append((s, e))
            s = e = x
    out.append((s, e))
    return out


def collect_bottom_border_commands(ws, mr, mc, min_span=6):
    cmds = []
    t_i = 0
    for r in range(1, mr + 1):
        for style in ("thick", "thin", "medium", "hair", "double"):
            cols = [c for c in range(1, mc + 1) if bottom_border_style(ws.cell(row=r, column=c)) == style]
            for c1, c2 in _consecutive_runs(cols):
                if c2 - c1 + 1 < min_span:
                    continue
                word = BOTTOM_STYLE_WORD.get(style, style)
                tmpl = BORDER_BOTTOM_TEMPLATES[t_i % len(BORDER_BOTTOM_TEMPLATES)]
                t_i += 1
                cmds.append(tmpl.format(r=r, c1=c1, c2=c2, style=word))
    return cmds


def collect_wrap_commands(ws, mr, mc):
    cmds = []
    t_i = 0
    for r in range(1, mr + 1):
        cols = []
        for c in range(1, mc + 1):
            a = ws.cell(row=r, column=c).alignment
            if a and a.wrap_text:
                cols.append(c)
        for c1, c2 in _consecutive_runs(cols):
            if c1 == c2:
                tmpl = WRAP_ONE_COL_TEMPLATES[t_i % len(WRAP_ONE_COL_TEMPLATES)]
                cmds.append(tmpl.format(c=c1, r=r))
            else:
                tmpl = WRAP_TEMPLATES[t_i % len(WRAP_TEMPLATES)]
                cmds.append(tmpl.format(c1=c1, c2=c2, r=r))
            t_i += 1
    return cmds


def describe_excel_number_format(nf):
    if not nf or nf == "General":
        return None
    if nf == "0%":
        return "a whole-number percentage"
    if nf == '"$"#,##0.00':
        return "currency with a dollar sign and two decimals"
    if nf.lower() in ("m/d/yy", "m/d/yyyy", "mm/dd/yy"):
        return "a short date (month, day, year)"
    return None


def collect_number_format_commands(ws, mr, mc):
    cmds = []
    t_i = 0
    for c in range(1, mc + 1):
        r = 1
        while r <= mr:
            nf = ws.cell(row=r, column=c).number_format
            if nf in (None, "General"):
                r += 1
                continue
            r2 = r
            while r2 <= mr and ws.cell(row=r2, column=c).number_format == nf:
                r2 += 1
            run = r2 - r
            desc = describe_excel_number_format(nf)
            if not desc:
                r = r2
                continue
            if run >= 5:
                tmpl = NUMFMT_RANGE_TEMPLATES[t_i % len(NUMFMT_RANGE_TEMPLATES)]
                cmds.append(tmpl.format(c=c, r1=r, r2=r2 - 1, desc=desc))
                t_i += 1
            elif nf.lower() in ("m/d/yy", "m/d/yyyy") and run >= 1:
                tmpl = NUMFMT_ONE_TEMPLATES[t_i % len(NUMFMT_ONE_TEMPLATES)]
                cmds.append(tmpl.format(c=c, r=r, desc=desc))
                t_i += 1
            r = r2
    return cmds


def collect_merge_commands(ws):
    ranges = sorted(ws.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col, m.max_row, m.max_col))
    cmds = []
    for i, m in enumerate(ranges):
        r1, r2, c1, c2 = m.min_row, m.max_row, m.min_col, m.max_col
        if r1 == r2:
            tmpl = MERGE_SINGLE_ROW_TEMPLATES[i % len(MERGE_SINGLE_ROW_TEMPLATES)]
            cmds.append(tmpl.format(r=r1, c1=c1, c2=c2))
        elif c1 == c2:
            tmpl = MERGE_MULTI_ONE_COL_TEMPLATES[i % len(MERGE_MULTI_ONE_COL_TEMPLATES)]
            cmds.append(tmpl.format(c=c1, r1=r1, r2=r2))
        else:
            tmpl = MERGE_MULTI_ROW_TEMPLATES[i % len(MERGE_MULTI_ROW_TEMPLATES)]
            cmds.append(tmpl.format(r1=r1, r2=r2, c1=c1, c2=c2))
    return cmds


def collect_uniform_row_height_command(ws, mr):
    heights = []
    for r in range(1, mr + 1):
        dim = ws.row_dimensions.get(r)
        h = dim.height if dim and dim.height is not None else None
        heights.append(h)
    if not heights or any(x is None for x in heights):
        return []
    ctr = Counter(heights)
    h_common, n = ctr.most_common(1)[0]
    if n == mr and h_common is not None:
        tmpl = ROW_HEIGHT_UNIFORM_TEMPLATES[0]
        return [tmpl.format(r1=1, r2=mr, h=int(h_common) if float(h_common).is_integer() else round(h_common, 1))]
    return []


def collect_column_width_commands(ws, mc, max_lines=14):
    rows = []
    for c in range(1, mc + 1):
        letter = get_column_letter(c)
        cd = ws.column_dimensions.get(letter)
        w = cd.width if cd and cd.width is not None else None
        if w is None:
            continue
        if w < 4.0 or w > 12.0:
            rows.append((c, w))
    rows.sort(key=lambda t: abs(t[1] - 8.5), reverse=True)
    cmds = []
    wi = ni = 0
    for c, w in rows[:max_lines]:
        rounded = round(w, 2)
        if w < 6.0:
            tmpl = COL_WIDTH_NARROW_TEMPLATES[ni % len(COL_WIDTH_NARROW_TEMPLATES)]
            ni += 1
        else:
            tmpl = COL_WIDTH_WIDE_TEMPLATES[wi % len(COL_WIDTH_WIDE_TEMPLATES)]
            wi += 1
        cmds.append(tmpl.format(c=c, w=rounded))
    return cmds


def collect_header_alignment_commands(ws, mc, max_header_row=3):
    cmds = []
    t_i = 0
    for r in range(1, max_header_row + 1):
        hcols = []
        for c in range(1, mc + 1):
            a = ws.cell(row=r, column=c).alignment
            if a and a.horizontal == "center":
                hcols.append(c)
        for c1, c2 in _consecutive_runs(hcols):
            if c2 - c1 + 1 < 2:
                continue
            tmpl = ALIGN_H_CENTER_TEMPLATES[t_i % len(ALIGN_H_CENTER_TEMPLATES)]
            cmds.append(tmpl.format(r=r, c1=c1, c2=c2))
            t_i += 1
    return cmds


def collect_headline_font_commands(ws, mc):
    cmds = []
    t_i = 0
    r = 1
    blue_cols = []
    for c in range(1, min(mc + 1, 40)):
        f = ws.cell(row=r, column=c).font
        if not f:
            continue
        sz = f.size
        fg = hx_from_color(getattr(f, "color", None))
        if sz and float(sz) >= 20 and fg and fg.upper() == "#0B5394":
            blue_cols.append((c, sz))
    for c1, c2 in _consecutive_runs([c for c, _ in blue_cols]):
        sz = next(s for cc, s in blue_cols if cc == c1)
        sz_s = int(sz) if float(sz).is_integer() else round(sz, 1)
        if c1 == c2:
            tmpl = HEADLINE_BLUE_TEMPLATES[t_i % len(HEADLINE_BLUE_TEMPLATES)]
            cmds.append(tmpl.format(c=c1, r=r, sz=sz_s))
        else:
            cmds.append(
                f"Use large blue headings (about {sz_s} point) for columns {c1} through {c2} on row {r}."
            )
        t_i += 1
    white_cols = []
    for c in range(1, mc + 1):
        f = ws.cell(row=r, column=c).font
        fg = hx_from_color(getattr(f, "color", None)) if f else None
        if fg and fg.upper() in ("#FFFFFF",):
            white_cols.append(c)
    for c1, c2 in _consecutive_runs(white_cols):
        if c2 - c1 + 1 < 3:
            continue
        tmpl = WHITE_FONT_TEMPLATES[t_i % len(WHITE_FONT_TEMPLATES)]
        cmds.append(tmpl.format(r=r, c1=c1, c2=c2))
        t_i += 1
    return cmds


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / "Downloads" / "Project timeline.xlsx"
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    ws = wb.active
    mr, mc = ws.max_row or 1, ws.max_column or 1

    bold_cmds = []
    fill_cmds = []

    for row in range(1, mr + 1):
        for col in range(1, mc + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            has_text = val is not None and str(val).strip() != ""
            font = cell.font
            bold = bool(font and font.bold)
            if bold and has_text:
                tmpl = BOLD_TEMPLATES[(row + col + len(bold_cmds)) % len(BOLD_TEMPLATES)]
                bold_cmds.append(tmpl.format(c=col, r=row))

            f = cell.fill
            pt = getattr(f, "patternType", None) or getattr(f, "fill_type", None)
            if pt not in ("solid", "solidFill"):
                continue
            hx = hx_from_color(getattr(f, "fgColor", None)) or hx_from_color(getattr(f, "bgColor", None))
            if not hx or hx in LIGHT_FILLS or hx == "#000000":
                continue
            color = nearest_color_word(hx)
            tmpl = FILL_TEMPLATES[(row + col + len(fill_cmds)) % len(FILL_TEMPLATES)]
            fill_cmds.append(tmpl.format(c=col, r=row, color=color))

    merge_cmds = collect_merge_commands(ws)
    row_height_cmds = collect_uniform_row_height_command(ws, mr)
    col_width_cmds = collect_column_width_commands(ws, mc)
    border_cmds = collect_bottom_border_commands(ws, mr, mc, min_span=6)
    wrap_cmds = collect_wrap_commands(ws, mr, mc)
    numfmt_cmds = collect_number_format_commands(ws, mr, mc)
    align_cmds = collect_header_alignment_commands(ws, mc)
    font_cmds = collect_headline_font_commands(ws, mc)

    for c in bold_cmds:
        print(c, flush=True)
    print("", flush=True)
    for c in fill_cmds:
        print(c, flush=True)
    print("", flush=True)
    for section in (
        merge_cmds,
        row_height_cmds,
        col_width_cmds,
        border_cmds,
        wrap_cmds,
        numfmt_cmds,
        align_cmds,
        font_cmds,
    ):
        for c in section:
            print(c, flush=True)
        if section:
            print("", flush=True)
    print(EXAMPLE_MULTI_ROW_RANGE, flush=True)


if __name__ == "__main__":
    main()
